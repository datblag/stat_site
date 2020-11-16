from openpyxl import Workbook, styles
import os, random, shutil, string, subprocess
import datetime
import logging

class TemplateRender:
    align_center_center = styles.Alignment(horizontal='center',
                                           vertical='center',
                                           text_rotation=0,
                                           wrap_text=True,
                                           shrink_to_fit=False,
                                           indent=0)

    align_center_center_not_wrap = styles.Alignment(horizontal='center',
                                                    vertical='center',
                                                    text_rotation=0,
                                                    wrap_text=False,
                                                    shrink_to_fit=False,
                                                    indent=0)


    align_left_top = styles.Alignment(horizontal='left',
                                           vertical='top',
                                           text_rotation=0,
                                           wrap_text=True,
                                           shrink_to_fit=False,
                                           indent=0)

    align_left_top_not_wrap = styles.Alignment(horizontal='left',
                                               vertical='top',
                                               text_rotation=0,
                                               wrap_text=False,
                                               shrink_to_fit=False,
                                               indent=0)

    def __init__(self, report_columns=1, header_width=50, file_name='', copyfile=False, open_in_excel=False,
                 dirs_copy_name=[], sheet_title=''):
        self.sheet_count = 1
        self.current_line = 0
        self.report_columns = report_columns
        self.header_width = header_width
        self.file_name = file_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.copyfile = copyfile
        self.open_in_excel = open_in_excel
        self.dirs_copy_name = dirs_copy_name
        self.first_data_row = None
        self.add_sheet(sheet_title)

    def add_sheet(self, sheet_title=''):
        self.current_line = 1
        if self.sheet_count > len(self.wb.worksheets):
            self.wb.create_sheet()
        self.ws = self.wb.worksheets[self.sheet_count-1]
        if sheet_title:
            self.ws.title = sheet_title
        self.sheet_count += 1

    def add_data_row(self, rows, aligment=None):
        for row in rows:
            for en, el in enumerate(row):
                if isinstance(el, datetime.date):
                    self.ws.cell(row=self.current_line, column=en + 1).value = el.strftime('%d.%m.%Y')
                else:
                    self.ws.cell(row=self.current_line, column=en + 1).value = el
                if aligment is not None:
                    self.ws.cell(row=self.current_line, column=en + 1).alignment = aligment
            if self.first_data_row is None:
                self.first_data_row = self.current_line
            self.current_line += 1

    def add_header_row(self, title):
        for col_num in range(1, self.report_columns+1):
            self.ws.cell(row=self.current_line, column=col_num).alignment = self.align_left_top_not_wrap
        self.ws.cell(row=self.current_line, column=1).value = title
        self.current_line += 1

    def add_titles_row(self, rows):
        for en, row in enumerate(rows):
            self.ws.cell(row=self.current_line, column=en + 1).value = row[0]
            self.ws.cell(row=self.current_line, column=en + 1).alignment = self.align_center_center

            self.ws.column_dimensions[self.ws.cell(row=self.current_line, column=en + 1).column_letter].width = row[1]

        self.current_line += 1

    def close_template_file(self):
        logging.warning(self.file_name)
        self.wb.save(filename=self.file_name)
        self.wb.close()
        # if self.copyfile:
        #     self.copy_template_file()
        if self.open_in_excel:
            self.open_template_in_excel()

    def copy_template_file(self):
        for dir_copy_name in self.dirs_copy_name:
            shutil.copy(self.file_name, os.path.join(dir_copy_name, self.file_name))

    def open_template_in_excel(self):
        temp_file_name = os.path.join(os.environ['TMP'],
                                      "report_" + ''.join(random.choice(string.ascii_letters +
                                                                        string.digits) for _ in range(10)) + '.xlsx')
        shutil.copy(self.file_name, temp_file_name)
        subprocess.Popen(temp_file_name, shell=True)

