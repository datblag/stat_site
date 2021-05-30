import openpyxl
from openpyxl.utils import column_index_from_string
import os
import dbf
import datetime
from site_app.models.mis_db import HltMkabTable, session_mis
from sqlalchemy import func
import logging
MAPPING_BY_DATE = 1
MAPPING_BY_YEAR = 0


def process_gastro_file(path_name, file_name, start_date=None, end_date=None, start_row_num=2, process_sheet_num=1):
    start_row = start_row_num - 1
    examinations = []
    if process_sheet_num == 1:
        service_code = 1844
    elif process_sheet_num == 2:
        service_code = 1846
    else:
        logging.error('Для указанного листа не задана услуга')
        return examinations
    sheet_num = process_sheet_num - 1
    file_full_name = os.path.join(path_name, file_name)
    wb = openpyxl.load_workbook(file_full_name)
    ws = wb.worksheets[sheet_num]
    import_total = 0

    logging.warning(end_date)
    for row_num, row in enumerate(ws.iter_rows()):
        if row_num < start_row:
            continue
        row_num += 1
        fio_value = row[0].value
        logging.warning(fio_value)
        fio_value = fio_value.split(' ')
        fam_value = ''
        im_value = ''
        ot_value = ''
        logging.warning(fio_value)
        exam = {'fam': fam_value, 'im': im_value, 'ot': ot_value, 'dr': row[1].value,  'd_u': row[2].value}
        # logging.warning(exam)
        if exam['fam'] is None or exam['fam'] =='':
            continue
        if exam['im'] is None:
            continue
        dr_value = exam['dr'][0:8]
        if len(dr_value) == 4:
            logging.warning(dr_value)
            dr_value = '01.01.' + dr_value
            logging.warning(dr_value)
        elif len(dr_value) == 6:
            dr_value = '01.01.' + dr_value[2:4]
        try:
            dr = datetime.datetime.strptime(dr_value, '%d.%m.%y')
        except:
            continue
        if dr > datetime.datetime.now():
            dr_value = dr_value[0:6] + '19' + dr_value[6:8]
            dr = datetime.datetime.strptime(dr_value, '%d.%m.%Y')
        exam['dr'] = dr

        du_value = exam['d_u'][0:8]
        du = datetime.datetime.strptime(du_value, '%d.%m.%y')
        exam['d_u'] = du
        logging.warning(exam)
        if start_date is not None and exam['d_u'] < start_date:
            continue
        if end_date is not None and exam['d_u'] > end_date:
            continue
        import_total += 1

    logging.warning('Всего импортировано строк ' + str(import_total))


def process_uzi_file(path_name, file_name, start_date=None, end_date=None, start_row_num=2, process_sheet_num=1):
    start_row = start_row_num - 1
    examinations = []
    if process_sheet_num == 1:
        service_code = 1844
    elif process_sheet_num == 2:
        service_code = 1846
    else:
        logging.error('Для указанного листа не задана услуга')
        return examinations
    sheet_num = process_sheet_num - 1
    file_full_name = os.path.join(path_name, file_name)
    wb = openpyxl.load_workbook(file_full_name)
    ws = wb.worksheets[sheet_num]
    import_total = 0

    logging.warning(end_date)
    for row_num, row in enumerate(ws.iter_rows()):
        if row_num < start_row:
            continue
        row_num += 1
        exam = {'fam': row[0].value, 'im': row[1].value,  'ot': row[2].value, 'dr': row[3].value, 'd_u': row[4].value,
                'cod': service_code, 'doctor': row[5].value}
        if exam['fam'] is None:
            continue
        if start_date is not None and exam['d_u'] < start_date:
            continue
        if end_date is not None and exam['d_u'] > end_date:
            continue

        # exam = [fam_value, im_value, ot_value, dr_value, d_u_value, service_code, doctor_value]
        # logging.warning(exam)
        examinations.append(exam)
        import_total += 1

    logging.warning('Всего импортировано строк ' + str(import_total))

    return examinations


def process_sogaz_control_file_to_csv(path_name, file_name):

    dbf_file = dbf.Table(r'S:\_out\reestr.dbf')
    dbf_file.open()

    index_numer = dbf_file.create_index(lambda rec: (rec.fam.strip().upper() + ' ' + rec.im.strip().upper()
                                                     + ' ' + rec.ot.strip().upper(), rec.dr, rec.ds.strip(), rec.cod,
                                                     rec.d_u))

    file_full_name = os.path.join(path_name, file_name)
    wb = openpyxl.load_workbook(file_full_name)
    start_row = 3 - 1
    ws = wb.worksheets[0]
    for row_num, row in enumerate(ws.iter_rows()):
        if row_num < start_row:
            continue
        row_num += 1
        fam_value = row[0].value
        if fam_value is None:
            continue
        if 'Итого' in fam_value:
            continue
        dr_value = row[1].value
        ds_value = row[4].value
        cod_value = row[6].value
        date_value1 = row[8].value
        date_value2 = ws.cell(row_num + 1, 9).value
        sum_value = row[10].value
        history_value = row[11].value
        recs = index_numer.search(match=(fam_value.strip(), datetime.datetime.strptime(dr_value, '%d.%m.%Y').date(),
                                         ds_value.strip(), int(cod_value),
                                         datetime.datetime.strptime(date_value2, '%d.%m.%Y').date()))
        # for rec_numer in recs:
        #     logging.warning(rec_numer)
        if len(recs) == 1:
            doctor_name = recs[0]['NAME']
            logging.warning(doctor_name)
            ws.cell(row_num, 15).value = doctor_name
        else:
            logging.warning([len(recs), fam_value, dr_value, ds_value, int(cod_value), date_value1, date_value2,
                             sum_value, history_value])

    dbf_file.close()
    wb.save(file_full_name)


def set_patient_num_to_excel(path_name, file_name, start_row_num, type_mapping_birthday, fam_col_index, dr_col_index,
                             result_col_index, worksheet_num=1, field_name='num'):
    # start_row = 2
    start_row = start_row_num - 1  # нумерация с 0 !!!!!!!!!!!!!!!!!!
    # type_mapping_birthday = MAPPING_BY_DATE

    if isinstance(fam_col_index, list):
        fam_col = column_index_from_string(fam_col_index[0])
        im_col = column_index_from_string(fam_col_index[1])
        ot_col = column_index_from_string(fam_col_index[2])
    else:
        fam_col = column_index_from_string(fam_col_index)
    dr_col = column_index_from_string(dr_col_index)
    result_col_history = column_index_from_string(result_col_index)

    find_records_count = 0

    file_full_name = os.path.join(path_name, file_name)
    wb = openpyxl.load_workbook(file_full_name)
    # ws = wb.active
    ws = wb.worksheets[worksheet_num-1]

    if ws.max_column <= result_col_history:
        ws.cell(1, result_col_history).value = ''

    for row_num, row in enumerate(ws.iter_rows()):
        if row_num < start_row:
            continue
        row_num += 1
        year_bird = row[dr_col-1].value
        fam = None
        im = None
        ot = None
        if isinstance(fam_col_index, list):
            fam = row[fam_col-1].value
            im = row[im_col-1].value
            ot = row[ot_col-1].value
            if fam is None:
                continue
            if im is None:
                continue
        else:
            fam_value = row[fam_col-1].value
            if fam_value is None:
                continue
            logging.warning(fam_value)
            fio_list = fam_value.strip().split(' ', 1)
            if len(fio_list) > 1:
                fam = fio_list[0]
                fam_value = fio_list[1]
                fio_list = fam_value.strip().split(' ', 1)
                im = fio_list[0]
                if len(fio_list) > 1:
                    ot = fio_list[1]
            else:
                continue

        logging.warning([fam, im, ot])

        query = session_mis.query(HltMkabTable.num.label('num'), HltMkabTable.ss.label('ss'))
        query = query.filter(func.rtrim(func.ltrim(HltMkabTable.family)) == fam.strip())
        query = query.filter(func.rtrim(func.ltrim(HltMkabTable.name)) == im.strip())
        if ot is not None:
            query = query.filter(func.rtrim(func.ltrim(HltMkabTable.ot)) == ot.strip())

        if type_mapping_birthday == MAPPING_BY_YEAR:
            query = query.filter(HltMkabTable.date_bd >= datetime.date(year=year_bird, month=1, day=1))
            query = query.filter(HltMkabTable.date_bd <= datetime.date(year=year_bird, month=12, day=31))
        elif type_mapping_birthday == MAPPING_BY_DATE:
            if isinstance(year_bird, str):
                year_bird = datetime.datetime.strptime(year_bird, '%d.%m.%Y')
            if isinstance(year_bird, datetime.datetime):
                query = query.filter(HltMkabTable.date_bd == year_bird)
            else:
                logging.error('Ошибка в дате рождения, поиск не выполнен')
                continue
        recs = query.all()
        logging.warning(recs)
        if recs:
            find_records_count += 1
            ws.cell(row_num, result_col_history).value = recs[0][recs[0]._fields.index(field_name)]
    logging.warning(find_records_count)
    wb.save(file_full_name)


def main():

    path_name = r'S:\_reestr\202105\obsl'
    file_name = 'узи омс июнь.xlsx'
    start_date = datetime.datetime(2021, 4, 23)
    end_date = datetime.datetime(2021, 5, 22)

    # разбор файла узи обсследований
    # process_sheet_num = 1  # 1-эхо, 2 - дуплекс
    # process_uzi_file(path_name, file_name, start_date, end_date, start_row_num=2, process_sheet_num=process_sheet_num)

    # разбор заявки согаза на экспертизу
    # process_sogaz_control_file_to_csv(path_name, file_name)

    # госпитализация гор больницы
    # set_patient_num_to_excel(path_name, file_name, start_row_num=10, type_mapping_birthday=MAPPING_BY_DATE,
    #                           fam_col_index='C', dr_col_index='D', result_col_index='T', worksheet_num=1,
    #                           field_name='num')

    # госпитализация областной больницы
    # set_patient_num_to_excel(path_name, file_name, start_row_num=2, type_mapping_birthday=MAPPING_BY_DATE,
    #                          fam_col_index='B', dr_col_index='C', result_col_index='K', worksheet_num=1,
    #                          field_name='num')

    # Обработка файла ФИО в обном столбце, дата рождения полностью в другом, проставить номера карт
    # set_patient_num_to_excel(path_name, file_name, start_row_num=3, type_mapping_birthday=MAPPING_BY_DATE,
    #                          fam_col_index='D', dr_col_index='F', result_col_index='I', worksheet_num=1,
    #                          field_name='num')


if __name__ == '__main__':
    main()

